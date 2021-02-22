import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter import font as tkfont
import pandas as pd
from openpyxl import load_workbook
from tkinter import messagebox
from datetime import date
import xlrd
from functools import partial 
from operator import ne

def getdata():
    global batch
    df = pd.read_excel('./data/studentlist.xlsx')
    df_rows = df.to_numpy().tolist()
    for value in df_rows:
        subdate1=value[22].split("/")
        if subdate1[2]!=batch:
            df_rows =list(filter(partial(ne, value), df_rows))
        else:
            continue
    return df_rows
def start():
    window3 = Tk()
    window3.title("Rivenburg Student Management System")
    window3.iconbitmap('data/icon.ico')
    width, height = window3.winfo_screenwidth(), window3.winfo_screenheight()
    window3.geometry('%dx%d+0+0' % (width,height))
    window3.state("zoomed")
    window3.configure(bg="white")
    bg = PhotoImage(file="data/bg.png")
    my_canvas = Canvas(window3, width=800, height=500)
    my_canvas.pack(fill="both", expand=True)
    my_canvas.create_image(0,0, image=bg, anchor="nw")
    txtfont=tkfont.Font(size=10)
    txtfont1=tkfont.Font(size=20,family="Times New Roman")
    txtfont2=tkfont.Font(size=14,family="Times New Roman")
    txtfont3=tkfont.Font(size=12)
    txtfont4=tkfont.Font(size=14)

    def adid():
        df = pd.read_excel('./data/studentlist.xlsx')
        df_rows = df.to_numpy().tolist()
        for value in df_rows:
            val=value[24]
        if val == "AdmissionID":
            newid="RVBS"+str(0).zfill(6)
        else:
            val1=val[4:]
            val2=int(val1)+1
            newid="RVBS"+str(val2).zfill(6)
        return newid
    def state1():
        global state
        global substate
        global batch
        global df_rows
        if substate==1:
            batch=today.strftime("%Y")
            df_rows=getdata()
            substate=0
        state=1
        my_canvas.delete("all")
        my_canvas.create_image(0,0, image=bg, anchor="nw")
        form()
    def state2():
        global state
        state=2
        my_canvas.delete("all")
        my_canvas.create_image(0,0, image=bg, anchor="nw")
        form()
    def state3():
        global state
        state=3
        my_canvas.delete("all")
        my_canvas.create_image(0,0, image=bg, anchor="nw")
        form()
    def state4():
        global state
        global substate
        tags=[]
        substate=1
        state=4
        my_canvas.delete("all")
        my_canvas.create_image(0,0, image=bg, anchor="nw")
        form()
    def state5():
        global state
        state=5
        my_canvas.delete("all")
        my_canvas.create_image(0,0, image=bg, anchor="nw")
        form()
    def form():
        global months
        global df_rows
        global batch
        months=[]
        v = StringVar()
        v.set("Tags: Students")
        u = StringVar()
        u.set("For the month of")
        v1 = StringVar()
        v1.set("Count: ")
        v2 = StringVar()
        v2.set("Tags: Entries")
        var1 = tk.IntVar()
        def check_fee(adid):
            feefile = pd.read_excel('./data/feestatus.xlsx')
            feedata = feefile.to_numpy().tolist()
            date=today.strftime("%d/%m/%Y")
            subdate=date.split("/")
            for value in feedata:
                if value[3]==adid:
                    for i in range(4,int(subdate[1])+3):
                        if value[i]=="not paid":
                            return 0
            return 1

        def addmonth():
            global months
            if variable.get()!="Month":
                months.append(variable.get())
                u.set(u.get()+", "+variable.get())
            return
        def to_excel():
            newid=adid()
            entry=[entry1.get(),entry2.get(),variable2.get(),entry4.get(),entry5.get(),entry6.get(),entry7.get(),entry8.get(),entry9.get(),entry10.get(),entry11.get(),entry12.get(),entry18.get(),entry13.get(),entry25.get(),variable.get(),entry15.get(),entry16.get(),entry17.get(),"0",entry19.get(),variable1.get(),today.strftime("%d/%m/%Y")," ",newid]
            for value in entry:
                if value=="":
                    messagebox.showwarning("Oops!","Some fields are still empty.")
                    return
            entryad=[entry1.get(),entry2.get(),newid,entry20.get(),today.strftime("%d/%m/%Y"),entry17.get()]
            for value in entryad:
                if value=="":
                    messagebox.showwarning("Oops!","Some fields are still empty.")
                    return
            wb = load_workbook('./data/studentlist.xlsx')
            ws = wb.worksheets[0]
            ws.append(entry)
            wb.save('./data/studentlist.xlsx')
            wb = load_workbook('./data/admissions.xlsx')
            ws = wb.worksheets[0]
            ws.append(entryad)
            wb.save('./data/admissions.xlsx')
            wb = load_workbook('./data/feestatus.xlsx')
            ws = wb.worksheets[0]
            entryfee=[entry1.get(),entry2.get(),entry17.get(),newid,"not paid","not paid","not paid","not paid","not paid","not paid","not paid","not paid","not paid","not paid","not paid","not paid",today.strftime("%d/%m/%Y")]
            ws.append(entryfee)
            wb.save('./data/feestatus.xlsx')
            messagebox.showinfo("Entry Sucessfull","Student data has been saved sucessfully.\nAdmission ID: "+newid)
            form()
        def to_excel1():
            global months
            entry=[entry1.get(),entry2.get(),entry3.get(),entry4.get(),entry5.get(),entry8.get(),entry6.get(),entry7.get()]
            for value in entry:
                if value=="" or variable.get()=="Month":
                    messagebox.showwarning("Oops!","Some fields are still empty.")
                    return
                else:
                    loc = ('data/studentlist.xlsx')
                    wb = xlrd.open_workbook(loc)
                    sheet = wb.sheet_by_index(0)
                    sheet.cell_value(0, 0)
                    for i in range(1,sheet.nrows):
                        if int(sheet.cell_value(i, 18))==int(entry2.get()):
                            if int(sheet.cell_value(i, 19))==int(entry3.get()):
                                adid=str(sheet.cell_value(i, 24))
                                year1=sheet.cell_value(i, 22).split('/')
                                year2=today.strftime("%Y")
                                if int(year1[2])==int(year2):
                                    wb = load_workbook('./data/studentlist.xlsx')
                                    ws = wb.worksheets[0]
                                    ws['X'+str(i+1)]=today.strftime("%d/%m/%Y")
                                    wb.save('./data/studentlist.xlsx')
                                    entry=[entry1.get(),entry2.get(),entry3.get(),entry4.get(),entry5.get(),entry8.get(),entry6.get(),entry7.get(),today.strftime("%d/%m/%Y")]
                                    wb = load_workbook('./data/fee.xlsx')
                                    ws = wb.worksheets[0]
                                    ws.append(entry)
                                    wb.save('./data/fee.xlsx')
                                    break
                    loc = ('data/feestatus.xlsx')
                    wb = xlrd.open_workbook(loc)
                    sheet = wb.sheet_by_index(0)
                    sheet.cell_value(0, 0)
                    for i in range(1,sheet.nrows):
                        if str(sheet.cell_value(i, 3))==adid:
                            for item in months:
                                    wb = load_workbook('./data/feestatus.xlsx')
                                    ws = wb.worksheets[0]
                                    ws[OptionList[item]+str(i+1)]=today.strftime("%d/%m/%Y")
                                    wb.save('./data/feestatus.xlsx')        
                            break
                        else:
                            if i==sheet.nrows:
                                messagebox.showwarning("Oops!","Student data not found..")
                                return
                            else:
                                continue   
                    messagebox.showinfo("Entry Sucessfull","Fee Entry has been saved sucessfully.")
                    break
            form()
        def to_excel2():
            tmp = pd.read_excel('./data/studentlist.xlsx')
            tmp_rows = tmp.to_numpy().tolist()
            entry=[entry1.get(),entry2.get(),entry4.get(),entry5.get(),today.strftime("%d/%m/%Y"),entry3.get()]
            for value in entry:
                if value=="":
                    messagebox.showwarning("Oops!","Some fields are still empty.")
                    return
                else:
                    loc = ('data/studentlist.xlsx')
                    wb = xlrd.open_workbook(loc)
                    sheet = wb.sheet_by_index(0)
                    sheet.cell_value(0, 0)
                    for i in range(1,sheet.nrows):
                        if str(sheet.cell_value(i, 24)).lower()==str(entry4.get()).lower():
                            wb = load_workbook('./data/studentlist.xlsx')
                            ws = wb.worksheets[0]
                            tmp_rows[i-1][22]=today.strftime("%d/%m/%Y")
                            tmp_rows[i-1][20]="Rivenburg School"
                            tmp_rows[i-1][21]=variable.get()
                            ws.append(tmp_rows[i-1])
                            wb.save('./data/studentlist.xlsx')
                            wb = load_workbook('./data/admissions.xlsx')
                            ws = wb.worksheets[0]
                            ws.append(entry)
                            wb.save('./data/admissions.xlsx')
                            break
                        else:
                            if i==sheet.nrows:
                                messagebox.showwarning("Oops!","Student data not found..")
                                return
                            else:
                                continue   
                    messagebox.showinfo("Entry Sucessfull","Fee Entry has been saved sucessfully.")
                    break
            form()
        def rd_excel():
            global df_rows
            global count
            global batch
            global tags
            count=0
            currentdate=date.today()
            if variable.get()=="Fee Status":
                if entry1.get()=="":
                    messagebox.showwarning("Oops!","Some fields are still empty.")
                    return
                else:
                    for value in df_rows:
                        if entry1.get().lower()=="paid":
                            if check_fee(value[24])== 1:
                                if var1.get()== 1:
                                    df_rows =list(filter(partial(ne, value), df_rows))
                                else:
                                    continue
                            else:
                                if var1.get()== 1:
                                    continue 
                                else:
                                    df_rows =list(filter(partial(ne, value), df_rows))
                        else:
                            messagebox.showwarning("Oops!","Invalid Entry.")
                            return
                    treeview.delete(*treeview.get_children())
                    for column in treeview["column"]:
                        treeview.heading(column, text=column,anchor="w",command=lambda _column=column: \
                                                         treeview_sort_column(treeview, _column, False))
                    for row in df_rows:
                        if count % 2 == 0:
                            treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('evenrow',))
                        else:
                            treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('oddrow',))
                        count=count+1
                    v1.set("Count: "+str(count))
                    if variable.get()+"="+entry1.get() not in tags:
                         tags.append("Fee Status="+entry1.get())
                    set_tags(var1.get())
                    entry1.delete(0,'end')
                    variable.set("Select Filter")
                    return

            elif variable.get()=="Batch":
                tags=[]
                if entry1.get()=="":
                    messagebox.showwarning("Oops!","Some fields are still empty.")
                    return
                else:
                    batch=entry1.get()
                    df_rows=getdata()
                    treeview.delete(*treeview.get_children())
                    for column in treeview["column"]:
                        treeview.heading(column, text=column,anchor="w",command=lambda _column=column: \
                                                         treeview_sort_column(treeview, _column, False))
                    for row in df_rows:
                        if count % 2 == 0:
                            treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('evenrow',))
                        else:
                            treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('oddrow',))
                        count=count+1
                    v1.set("Count: "+str(count))
                    if variable.get()+"="+entry1.get() not in tags:
                        tags.append("Batch="+entry1.get())
                    set_tags(var1.get())
                    entry1.delete(0,'end')
                    variable.set("Select Filter")
                    return
                    
            elif variable.get()!="Select Filter":
                if entry1.get()=="":
                    messagebox.showwarning("Oops!","Some fields are still empty.")
                    return
                else:
                    txt=entry1.get().lower().split(",")
                    tmp_rows=[]
                    for item in txt:
                        if var1.get()==1:
                            df_rows=df_rows
                        elif var1.get()==0 and ("," in entry1.get()):
                            df_rows = getdata()
                        for value in df_rows:
                            if str(value[OptionList[variable.get()]]).lower()!=item:
                                if var1.get()== 1:
                                    continue
                                else:
                                    df_rows =list(filter(partial(ne, value), df_rows))
                            else:
                                if var1.get()== 1:
                                    df_rows =list(filter(partial(ne, value), df_rows))
                                else:
                                    continue
                        if var1.get()==1:
                            tmp_rows=df_rows
                        elif var1.get()==0:
                            tmp_rows=tmp_rows+df_rows 
                    treeview.delete(*treeview.get_children())
                    for column in treeview["column"]:
                        treeview.heading(column, text=column,anchor="w",command=lambda _column=column: \
                                                         treeview_sort_column(treeview, _column, False))
                    for row in tmp_rows:
                        if count % 2 == 0:
                            treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('evenrow',))
                        else:
                            treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('oddrow',))
                        count=count+1
                    v1.set("Count: "+str(count))
                    if variable.get()+"="+entry1.get() not in tags:
                        tags.append(variable.get()+"="+entry1.get())
                    set_tags(var1.get())
                    entry1.delete(0,'end')
                    variable.set("Select Filter")
                    df_rows=tmp_rows
                    return

            else:
                messagebox.showwarning("Oops!","Please select a filter")
                entry1.delete(0,'end')
                return
            return
        def refresh_excel():
            global df_rows
            global count
            global batch
            count=0
            batch=today.strftime("%Y")
            df_rows = getdata()
            treeview.delete(*treeview.get_children())
            for column in treeview["column"]:
                treeview.heading(column, text=column,anchor="w",command=lambda _column=column: \
                                                 treeview_sort_column(treeview, _column, False))
            for row in df_rows:
                if count % 2 == 0:
                    treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('evenrow',))
                else:
                    treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('oddrow',))
                count=count+1
            v.set("Tags: Students")
            v1.set("Count: "+str(count))
            entry1.delete(0,'end')
            return
        def remove_filter():
            global df_rows
            global batch
            global tags
            entry1.delete(0,'end')
            batch=today.strftime("%Y")
            df_rows = getdata()
            v.set("Tags: Students")
            v1.set("Count: "+str(count))
            if tags:
                tags.pop()
            if not tags:
                refresh_excel()
            else:
                for items in tags:
                    data=items.split("=")
                    variable.set(data[0])
                    entry1.insert(0,data[1])
                    rd_excel() 
                    tags.pop()       
            return
        def treeview_sort_column(tv, col, reverse):
            l = [(tv.set(k, col), k) for k in tv.get_children('')]
            l.sort(reverse=reverse)

        # rearrange items in sorted positions
            for index, (val, k) in enumerate(l):
                tv.move(k, '', index)

        # reverse sort next time
            tv.heading(col, text=col, command=lambda _col=col: \
                         treeview_sort_column(tv, _col, not reverse))

        def get_entry(tmp_rows,index):
            global batch
            for value in tmp_rows:
                subdate1=value[index].split("/")
                if subdate1[2]!=batch:
                    tmp_rows =list(filter(partial(ne, value), tmp_rows))
                else:
                    continue
            return tmp_rows
        def set_tags(check):
            global tags
            global state
            if state==1:
                v.set("Tags: Students")
                for items in tags:
                    tag1=items.split("=")
                    if check==1:
                        v.set(v.get()+", "+tag1[0]+" not "+tag1[1])
                    else:
                        v.set(v.get()+", "+tag1[0]+" "+tag1[1])
            if state==4:
                v2.set("Tags: Entries")
                for items in tags:
                    tag1=items.split("=")
                    if check==1:
                        v2.set(v2.get()+", "+tag1[0]+" not "+tag1[1])
                    else:
                        v2.set(v2.get()+", "+tag1[0]+" "+tag1[1])
            return                    

        def log(check):
            def refresh():
                menu1['menu'].delete(0, 'end')
                for choice in OptionList1:
                    menu1['menu'].add_command(label=choice, command=tk._setit(variable1, choice))
                return
            global index
            global df_rows
            global cls
            global count
            global tags
            global batch
            i=0
            feeamnt=0
            addamnt=0
            basamnt=0
            uniamnt=0
            othamnt=0
            v1.set("Count: ")
            col=["","Count","Amount","","","",""]
            treeview.delete(*treeview.get_children())
            if variable.get()=="Total":
                if check==1:
                    tags=[]
                menu1['menu'].delete(0, 'end')
                for choice in OptionList1:
                    if i<2:
                        menu1['menu'].add_command(label=choice, command=tk._setit(variable1, choice))
                    i=i+1
                i=0
                index=0
                treeview["column"] = col
                treeview["show"] = "headings"
                treeview.tag_configure('oddrow', background="white")
                treeview.tag_configure('evenrow', background="#F7F7F7")
                for column in treeview["column"]:
                    treeview.heading(column, text=column,anchor="w",command=lambda _column=column: \
                                                     treeview_sort_column(treeview, _column, False))

                if entry1.get()!="" and entry2.get()!="":
                    df = pd.read_excel('./data/admissions.xlsx')
                    df_rows = df.to_numpy().tolist()
                    for value in df_rows:
                        date1=value[4]
                        subdate1=date1.split("/")
                        date2=entry1.get()
                        subdate2=date2.split("/")
                        date3=entry2.get()
                        subdate3=date3.split("/")
                        if (int(subdate1[2])>int(subdate2[2])) or ((int(subdate1[2])==int(subdate2[2])) and (int(subdate1[1])>int(subdate2[1]))) or ((int(subdate1[2])==int(subdate2[2])) and (int(subdate1[1])==int(subdate2[1])) and (int(subdate1[0])>=int(subdate2[0]))):
                            if (int(subdate1[2])<int(subdate3[2])) or ((int(subdate1[2])==int(subdate3[2])) and (int(subdate1[1])<int(subdate3[1]))) or ((int(subdate1[2])==int(subdate3[2])) and (int(subdate1[1])==int(subdate3[1])) and (int(subdate1[0])<=int(subdate3[0]))):
                                if var1.get()==1:
                                   if (int(value[5]) not in cls or (not cls)):
                                        i=i+1
                                        addamnt=addamnt+int(value[3])
                                else:
                                    if (int(value[5]) in cls or (not cls)):
                                        i=i+1
                                        addamnt=addamnt+int(value[3])
                                
                    addcount=i
                    i=0;
                    df = pd.read_excel('./data/fee.xlsx')
                    df_rows = df.to_numpy().tolist()
                    for value in df_rows:
                        date1=value[8]
                        subdate1=date1.split("/")
                        date2=entry1.get()
                        subdate2=date2.split("/")
                        date3=entry2.get()
                        subdate3=date3.split("/")
                        if (int(subdate1[2])>int(subdate2[2])) or ((int(subdate1[2])==int(subdate2[2])) and (int(subdate1[1])>int(subdate2[1]))) or ((int(subdate1[2])==int(subdate2[2])) and (int(subdate1[1])==int(subdate2[1])) and (int(subdate1[0])>=int(subdate2[0]))):
                            if (int(subdate1[2])<int(subdate3[2])) or ((int(subdate1[2])==int(subdate3[2])) and (int(subdate1[1])<int(subdate3[1]))) or ((int(subdate1[2])==int(subdate3[2])) and (int(subdate1[1])==int(subdate3[1])) and (int(subdate1[0])<=int(subdate3[0]))):
                                if var1.get()==1:
                                    if (int(value[1]) not in cls or (not cls)):
                                        i=i+1
                                        feeamnt=feeamnt+int(value[3])
                                        basamnt=basamnt+int(value[4])
                                        uniamnt=uniamnt+int(value[5])
                                        othamnt=othamnt+int(value[6])   
                                else:
                                    if (int(value[1])in cls or (not cls)):
                                        i=i+1
                                        feeamnt=feeamnt+int(value[3])
                                        basamnt=basamnt+int(value[4])
                                        uniamnt=uniamnt+int(value[5])
                                        othamnt=othamnt+int(value[6])
                                 
                            
                    feecount=i
                else:
                    df = pd.read_excel('./data/admissions.xlsx')
                    df_rows = df.to_numpy().tolist()
                    df_rows=get_entry(df_rows,4)
                    for value in df_rows:
                        if var1.get()==1:
                            if (int(value[5])not in cls or (not cls)):
                                i=i+1
                                addamnt=addamnt+int(value[3])
                        else:
                            if (int(value[5])in cls or (not cls)):
                                i=i+1
                                addamnt=addamnt+int(value[3])
                    addcount=i
                    i=0
                    df = pd.read_excel('./data/fee.xlsx')
                    df_rows = df.to_numpy().tolist()
                    df_rows=get_entry(df_rows,8)
                    for value in df_rows:
                        if var1.get()==1:
                            if (int(value[1])not in cls or (not cls)):
                                i=i+1
                                feeamnt=feeamnt+int(value[3])
                                basamnt=basamnt+int(value[4])
                                uniamnt=uniamnt+int(value[5])
                                othamnt=othamnt+int(value[6])
                        else:
                            if (int(value[1])in cls or (not cls)):
                                i=i+1
                                feeamnt=feeamnt+int(value[3])
                                basamnt=basamnt+int(value[4])
                                uniamnt=uniamnt+int(value[5])
                                othamnt=othamnt+int(value[6])

                    feecount=i
                count=0
                data=[["Admissions",addcount,addamnt],["Fee Entries",feecount,feeamnt],["Books And Stationaries","NA",basamnt],["Uniform","NA",uniamnt],["Other","NA",othamnt]]
                for row in data:
                    if count % 2 == 0:
                        treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('evenrow',))
                    else:
                        treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('oddrow',))
                    count=count+1
                cls=[]
                return

                    
            elif variable.get()=="Admissions":
                if check==1:
                    tags=[]
                menu1['menu'].delete(0, 'end')
                for choice in OptionList1:
                        menu1['menu'].add_command(label=choice, command=tk._setit(variable1, choice))
                if "Roll No." in OptionList1:
                    OptionList1["AdmissionID"] = OptionList1.pop("Roll No.")
                    refresh()
                index=4
                OptionList1["Class"]=5
                OptionList1["AdmissionID"]=2
                df = pd.read_excel('./data/admissions.xlsx')
                treeview["column"] = list(df.columns)
                treeview["show"] = "headings"
                treeview.tag_configure('oddrow', background="white")
                treeview.tag_configure('evenrow', background="#F7F7F7")
                for column in treeview["column"]:
                    treeview.heading(column, text=column,anchor="w",command=lambda _column=column: \
                                                     treeview_sort_column(treeview, _column, False))

                if entry1.get()!="" and entry2.get()!="":
                    df = pd.read_excel('./data/admissions.xlsx')
                    df_rows = df.to_numpy().tolist()
                    for value in df_rows:
                        date1=value[4]
                        subdate1=date1.split("/")
                        date2=entry1.get()
                        subdate2=date2.split("/")
                        date3=entry2.get()
                        subdate3=date3.split("/")
                        if (int(subdate1[2])>int(subdate2[2])) or ((int(subdate1[2])==int(subdate2[2])) and (int(subdate1[1])>int(subdate2[1]))) or ((int(subdate1[2])==int(subdate2[2])) and (int(subdate1[1])==int(subdate2[1])) and (int(subdate1[0])>=int(subdate2[0]))):
                            if (int(subdate1[2])<int(subdate3[2])) or ((int(subdate1[2])==int(subdate3[2])) and (int(subdate1[1])<int(subdate3[1]))) or ((int(subdate1[2])==int(subdate3[2])) and (int(subdate1[1])==int(subdate3[1])) and (int(subdate1[0])<=int(subdate3[0]))):
                                continue
                            else:
                                df_rows =list(filter(partial(ne, value), df_rows))
                        else:
                            df_rows =list(filter(partial(ne, value), df_rows))
                else:
                    df = pd.read_excel('./data/admissions.xlsx')
                    df_rows = df.to_numpy().tolist()
                    df_rows=get_entry(df_rows,index)
                count=0
                for row in df_rows:
                    if count % 2 == 0:
                        treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('evenrow',))
                    else:
                        treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('oddrow',))
                    count=count+1
                v1.set("Count: "+str(count))
                return
            elif variable.get()=="Fee Entries":
                if check==1:
                    tags=[]
                index=8
                menu1['menu'].delete(0, 'end')
                for choice in OptionList1:
                        menu1['menu'].add_command(label=choice, command=tk._setit(variable1, choice))
                if "AdmissionID" in OptionList1:
                    OptionList1["Roll No."] = OptionList1.pop("AdmissionID")
                    refresh()
                
                OptionList1["Class"]=1
                OptionList1["Roll No."]=2
                df = pd.read_excel('./data/fee.xlsx')
                treeview["column"] = list(df.columns)
                treeview["show"] = "headings"
                treeview.tag_configure('oddrow', background="white")
                treeview.tag_configure('evenrow', background="#F7F7F7")
                for column in treeview["column"]:
                    treeview.heading(column, text=column,anchor="w",command=lambda _column=column: \
                                                     treeview_sort_column(treeview, _column, False))

                if entry1.get()!="" and entry2.get()!="":
                    df = pd.read_excel('./data/fee.xlsx')
                    df_rows = df.to_numpy().tolist()
                    for value in df_rows:
                        date1=value[8]
                        subdate1=date1.split("/")
                        date2=entry1.get()
                        subdate2=date2.split("/")
                        date3=entry2.get()
                        subdate3=date3.split("/")
                        if (int(subdate1[2])>int(subdate2[2])) or ((int(subdate1[2])==int(subdate2[2])) and (int(subdate1[1])>int(subdate2[1]))) or ((int(subdate1[2])==int(subdate2[2])) and (int(subdate1[1])==int(subdate2[1])) and (int(subdate1[0])>=int(subdate2[0]))):
                            if (int(subdate1[2])<int(subdate3[2])) or ((int(subdate1[2])==int(subdate3[2])) and (int(subdate1[1])<int(subdate3[1]))) or ((int(subdate1[2])==int(subdate3[2])) and (int(subdate1[1])==int(subdate3[1])) and (int(subdate1[0])<=int(subdate3[0]))):
                                continue
                            else:
                                df_rows =list(filter(partial(ne, value), df_rows))
                        else:
                            df_rows =list(filter(partial(ne, value), df_rows))
                else:
                    df = pd.read_excel('./data/fee.xlsx')
                    df_rows = df.to_numpy().tolist()
                    df_rows=get_entry(df_rows,index)
                for row in df_rows:
                    if count % 2 == 0:
                        treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('evenrow',))
                    else:
                        treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('oddrow',))
                    count=count+1
                count=0
                v1.set("Count: "+str(count))
                return
            elif variable.get()=="Fee Status":
                if check==1:
                    tags=[]
                menu1['menu'].delete(0, 'end')
                for choice in OptionList1:
                        menu1['menu'].add_command(label=choice, command=tk._setit(variable1, choice))
                if "Roll No." in OptionList1:
                    OptionList1["AdmissionID"] = OptionList1.pop("Roll No.")
                    refresh()
                index=16
                OptionList1["Class"]=2
                OptionList1["AdmissionID"]=3
                df = pd.read_excel('./data/feestatus.xlsx')
                treeview["column"] = list(df.columns)
                treeview["show"] = "headings"
                treeview.tag_configure('oddrow', background="white")
                treeview.tag_configure('evenrow', background="#F7F7F7")
                for column in treeview["column"]:
                    treeview.heading(column, text=column,anchor="w",command=lambda _column=column: \
                                                     treeview_sort_column(treeview, _column, False))

                if entry1.get()!="" and entry2.get()!="":
                    df = pd.read_excel('./data/feestatus.xlsx')
                    df_rows = df.to_numpy().tolist()
                    for value in df_rows:
                        date1=value[16]
                        subdate1=date1.split("/")
                        date2=entry1.get()
                        subdate2=date2.split("/")
                        date3=entry2.get()
                        subdate3=date3.split("/")
                        if (int(subdate1[2])>int(subdate2[2])) or ((int(subdate1[2])==int(subdate2[2])) and (int(subdate1[1])>int(subdate2[1]))) or ((int(subdate1[2])==int(subdate2[2])) and (int(subdate1[1])==int(subdate2[1])) and (int(subdate1[0])>=int(subdate2[0]))):
                            if (int(subdate1[2])<int(subdate3[2])) or ((int(subdate1[2])==int(subdate3[2])) and (int(subdate1[1])<int(subdate3[1]))) or ((int(subdate1[2])==int(subdate3[2])) and (int(subdate1[1])==int(subdate3[1])) and (int(subdate1[0])<=int(subdate3[0]))):
                                continue
                            else:
                                df_rows =list(filter(partial(ne, value), df_rows))
                        else:
                            df_rows =list(filter(partial(ne, value), df_rows))
                else:
                    df = pd.read_excel('./data/feestatus.xlsx')
                    df_rows = df.to_numpy().tolist()
                    df_rows=get_entry(df_rows,index)
                count=0
                for row in df_rows:
                    if count % 2 == 0:
                        treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('evenrow',))
                    else:
                        treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('oddrow',))
                    count=count+1
                v1.set("Count: "+str(count))
                return

                    
        def excel_read():
            global df_rows
            global count
            global index
            global batch
            global cls
            global tags
            count=0
            if entry3.get()=="":
                messagebox.showwarning("Oops!","Some fields are still empty.")
                return
            elif variable1.get()=="Select Filter":
                messagebox.showwarning("Oops!","Please Select a filter.")
                return
            if variable1.get()=="Batch":
                entry1.delete(0,'end')
                entry2.delete(0,'end')
                entry1.insert(0,"1/1/"+entry3.get())
                entry2.insert(0,"31/12/"+entry3.get())
                batch=entry3.get()
                if variable1.get()+"="+entry3.get() not in tags:
                    tags.append(variable1.get()+"="+entry3.get())
                
                if variable1.get()!="Select Filter" and entry3.get()!="":
                    set_tags(var1.get())
                entry3.delete(0,'end')
                variable1.set("Select Filter")
                log(0)
                return


            elif variable.get()=="Total" and variable1.get()=="Class":
                items=entry3.get().split(',')
                for item in items:
                    cls.append(int(item))
                if variable1.get()+"="+entry3.get() not in tags:
                    tags.append(variable1.get()+"="+entry3.get())
                if variable1.get()!="Select Filter" and entry3.get()!="":
                    set_tags(var1.get())
                entry3.delete(0,'end')
                variable1.set("Select Filter")
                log(0)
                return

            elif variable1.get()!="Select Filter":
                if entry3.get()=="":
                    messagebox.showwarning("Oops!","Some fields are still empty.")
                    return
                else:
                    txt=entry3.get().lower().split(",")
                    tmp_rows=[]
                    tmp_rows1=df_rows
                    for item in txt:
                        if var1.get()==1:
                            tmp_rows1=tmp_rows1
                        elif var1.get()==0 and ("," in entry3.get()):
                            tmp_rows1 = get_entry(df_rows,index)
                        for value in tmp_rows1:
                            if str(value[OptionList1[variable1.get()]]).lower()!=item:
                                if var1.get()== 1:
                                    continue
                                else:
                                    tmp_rows1=list(filter(partial(ne, value), tmp_rows1))
                            else:
                                if var1.get()== 1:
                                    tmp_rows1=list(filter(partial(ne, value), tmp_rows1))
                                else:
                                    continue
                        if var1.get()==1:
                            tmp_rows=tmp_rows1
                        elif var1.get()==0:
                            tmp_rows=tmp_rows+tmp_rows1 
                    treeview.delete(*treeview.get_children())
                    for column in treeview["column"]:
                        treeview.heading(column, text=column,anchor="w",command=lambda _column=column: \
                                                         treeview_sort_column(treeview, _column, False))
                    count=0
                    for row in tmp_rows:
                        if count % 2 == 0:
                            treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('evenrow',))
                        else:
                            treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('oddrow',))
                        count=count+1
                    v1.set("Count: "+str(count))
                    if variable1.get()+"="+entry3.get() not in tags:
                        tags.append(variable1.get()+"="+entry3.get())
                    set_tags(var1.get())
                    entry3.delete(0,'end')
                    variable1.set("Select Filter")
                    df_rows=tmp_rows
                    
                    return
        def remove_filter1():
            global df_rows
            global batch
            global tags
            global index
            entry3.delete(0,'end')
            entry1.delete(0,'end')
            entry2.delete(0,'end')
            log(0)
            v2.set("Tags: Entries")
            v1.set("Count: "+str(count))
            if tags:
                tags.pop()
            if not tags:
                batch=today.strftime("%Y")
                log(1)
            else:
                for items in tags:
                    data=items.split("=")
                    variable1.set(data[0])
                    entry3.insert(0,data[1])
                    excel_read()
            return
  
        def OnDoubleClick(event):
            def update_det():
                global df_rows
                global batch
                if entry1.get()=="":
                    messagebox.showwarning("Oops!","Some fields are still empty.")
                    return
                else:
                    loc = ('data/studentlist.xlsx')
                    wb = xlrd.open_workbook(loc)
                    sheet = wb.sheet_by_index(0)
                    sheet.cell_value(0, 0)
                    for i in range(1,sheet.nrows):
                        if sheet.cell_value(i, 24)==val[24]:
                            wb = load_workbook('./data/studentlist.xlsx')
                            ws = wb.worksheets[0]
                            ws[str(OptionList[variable.get()])+str(i+1)]=str(entry1.get())
                            wb.save('./data/studentlist.xlsx')
                            break
                        else:
                            if i==sheet.nrows:
                                messagebox.showwarning("Oops!","Student data not found..")
                                return
                            else:
                                continue   
                    messagebox.showinfo("Entry Sucessfull","Student Data has been saved sucessfully.")
                    batch=today.strftime("%Y")
                    df_rows=getdata()
                    state1()
                
                window2.destroy()
            window2 = Tk()
            window2.title("Update Student Details")
            window2.iconbitmap('data/icon.ico')
            w = 285 # width for the Tk root
            h = 100 # height for the Tk root
            ws = window2.winfo_screenwidth() # width of the screen
            hs = window2.winfo_screenheight() # height of the screen
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            window2.geometry('%dx%d+%d+%d' % (w, h, x, y))
            window2.resizable(width=False, height=False)
            my_canvas = Canvas(window2, width=285, height=100)
            my_canvas.pack(fill="both", expand=True)
            OptionList={"Class":"S","Roll No.":"T","First Name":"A","Last name":"B","Gender":"C","Age":"D","Father's Name":"E","Mother's Name":"F","Gaurdian's Name":"G","DOB":"H","Religion":"I","Address":"J","Father's Contact No.":"K","Mother's Contact No.":"L","Identity Mark":"M","Blood Group":"N","Community":"O","Tribe/Caste":"P"}
            variable = StringVar(window2)
            menu=OptionMenu(window2,variable,*OptionList)
            menu.config(width=15,anchor='w')
            menu_window = my_canvas.create_window(5, 10, anchor="nw", window=menu)
            variable.set("Roll No.")
            entry1=Entry(window2,width=20)
            entry1_window = my_canvas.create_window(150, 15, anchor="nw", window=entry1)
            button4 = Button(window2, text="SAVE", command=update_det)
            button4.config(height = 1, width=10)
            button4_window = my_canvas.create_window(190, 50, anchor="nw", window=button4)
            item = treeview.selection()
            val=treeview.item(item)["values"]

        def log_call():
            v2.set("Tags: Entries")
            log(1)

            
        if state == 2:
            OptionList=['ST','SC','OBC','General']
            variable = StringVar(window3)
            variable.set("General")
            OptionList1=['Pass','Fail']
            variable1 = StringVar(window3)
            variable1.set("Pass")
            OptionList2=['Male','Female','Other']
            variable2 = StringVar(window3)
            variable2.set("Male")
            button5 = Button(window3, text="Student List", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state1)
            button5.config(height = 2, width=30)
            button5_window = my_canvas.create_window(2, 298, anchor="nw", window=button5)
            button1 = Button(window3, text="New Admission", bg='#ffffff', highlightthickness=0, bd=0, fg='#3f51cd',font=txtfont)
            button1.config(height = 2, width=30)
            button1_window = my_canvas.create_window(2, 338, anchor="nw", window=button1)
            button6 = Button(window3, text="Renew Admission", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=state5)
            button6.config(height = 2, width=30)
            button6_window = my_canvas.create_window(2, 378, anchor="nw", window=button6)
            button2 = Button(window3, text="Fee Entry", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state3)
            button2.config(height = 2, width=30)
            button2_window = my_canvas.create_window(2, 418, anchor="nw", window=button2)
            button3 = Button(window3, text="Entry Log", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state4)
            button3.config(height = 2, width=30)
            button3_window = my_canvas.create_window(2, 458, anchor="nw", window=button3)


            section1=Label(text="New Admission",font=txtfont1,bg='#ffffff')
            section1_window = my_canvas.create_window(350, 80, anchor="nw", window=section1)

            field1=Label(text="First Name",font=txtfont2,bg='#ffffff')
            field1_window = my_canvas.create_window(366, 160, anchor="nw", window=field1)
            entry1=Entry(window3,width=20,font=txtfont3)
            entry1_window = my_canvas.create_window(370, 200, anchor="nw", window=entry1)

            field2=Label(text="Last Name",font=txtfont2,bg='#ffffff')
            field2_window = my_canvas.create_window(366, 260, anchor="nw", window=field2)
            entry2=Entry(window3,width=20,font=txtfont3)
            entry2_window = my_canvas.create_window(370, 300, anchor="nw", window=entry2)

            field3=Label(text="Gender",font=txtfont2,bg='#ffffff')
            field3_window = my_canvas.create_window(366, 360, anchor="nw", window=field3)
            menu2=OptionMenu(window3,variable2,*OptionList2)
            menu2.config(font=txtfont,bg='#ffffff',width=17,anchor='w')
            menu_window2 = my_canvas.create_window(370, 400, anchor="nw", window=menu2)

            field4=Label(text="Age",font=txtfont2,bg='#ffffff')
            field4_window = my_canvas.create_window(366, 460, anchor="nw", window=field4)
            entry4=Lotfi(window3,width=20,font=txtfont3)
            entry4_window = my_canvas.create_window(370, 500, anchor="nw", window=entry4)

            field5=Label(text="Father's Name",font=txtfont2,bg='#ffffff')
            field5_window = my_canvas.create_window(366, 560, anchor="nw", window=field5)
            entry5=Entry(window3,width=20,font=txtfont3)
            entry5_window = my_canvas.create_window(370, 600, anchor="nw", window=entry5)

            field6=Label(text="Mother's Name",font=txtfont2,bg='#ffffff')
            field6_window = my_canvas.create_window(366, 660, anchor="nw", window=field6)
            entry6=Entry(window3,width=20,font=txtfont3)
            entry6_window = my_canvas.create_window(370, 700, anchor="nw", window=entry6)




            field7=Label(text="Gaurdian's Name",font=txtfont2,bg='#ffffff')
            field7_window = my_canvas.create_window(666, 160, anchor="nw", window=field7)
            entry7=Entry(width=20,font=txtfont3)
            entry7_window = my_canvas.create_window(670, 200, anchor="nw", window=entry7)

            field8=Label(text="Date Of Birth (dd/mm/yy)",font=txtfont2,bg='#ffffff')
            field8_window = my_canvas.create_window(666, 260, anchor="nw", window=field8)
            entry8=Entry(window3,width=20,font=txtfont3)
            entry8_window = my_canvas.create_window(670, 300, anchor="nw", window=entry8)

            field9=Label(text="Religion",font=txtfont2,bg='#ffffff')
            field9_window = my_canvas.create_window(666, 360, anchor="nw", window=field9)
            entry9=Entry(window3,width=20,font=txtfont3)
            entry9_window = my_canvas.create_window(670, 400, anchor="nw", window=entry9)

            field10=Label(text="Address",font=txtfont2,bg='#ffffff')
            field10_window = my_canvas.create_window(666, 460, anchor="nw", window=field10)
            entry10=Entry(window3,width=20,font=txtfont3)
            entry10_window = my_canvas.create_window(670, 500, anchor="nw", window=entry10)

            field11=Label(text="Father's Contact No.",font=txtfont2,bg='#ffffff')
            field11_window = my_canvas.create_window(666, 560, anchor="nw", window=field11)
            entry11=Lotfi(window3,width=20,font=txtfont3)
            entry11_window = my_canvas.create_window(670, 600, anchor="nw", window=entry11)

            field12=Label(text="Mother's Contact No.",font=txtfont2,bg='#ffffff')
            field12_window = my_canvas.create_window(666, 660, anchor="nw", window=field12)
            entry12=Lotfi(window3,width=20,font=txtfont3)
            entry12_window = my_canvas.create_window(670, 700, anchor="nw", window=entry12)




            field18=Label(text="Idendity Mark",font=txtfont2,bg='#ffffff')
            field18_window = my_canvas.create_window(966, 160, anchor="nw", window=field18)
            entry18=Entry(window3,width=20,font=txtfont3)
            entry18_window = my_canvas.create_window(970, 200, anchor="nw", window=entry18)

            field13=Label(text="Blood Group",font=txtfont2,bg='#ffffff')
            field13_window = my_canvas.create_window(966, 260, anchor="nw", window=field13)
            entry13=Entry(window3,width=20,font=txtfont3)
            entry13_window = my_canvas.create_window(970, 300, anchor="nw", window=entry13)

            field14=Label(text="Tribe/Caste",font=txtfont2,bg='#ffffff')
            field14_window = my_canvas.create_window(966, 360, anchor="nw", window=field14)

            menu=OptionMenu(window3,variable,*OptionList)
            menu.config(font=txtfont,bg='#ffffff',width=17,anchor='w')
            menu_window = my_canvas.create_window(970, 400, anchor="nw", window=menu)


            field25=Label(text="Community",font=txtfont2,bg='#ffffff')
            field25_window = my_canvas.create_window(966, 460, anchor="nw", window=field25)
            entry25=Entry(window3,width=20,font=txtfont3)
            entry25_window = my_canvas.create_window(970, 500, anchor="nw", window=entry25)
            
            field15=Label(text="Father's Occupation",font=txtfont2,bg='#ffffff')
            field15_window = my_canvas.create_window(966, 560, anchor="nw", window=field15)
            entry15=Entry(window3,width=20,font=txtfont3)
            entry15_window = my_canvas.create_window(970, 600, anchor="nw", window=entry15)

            field16=Label(text="Mother's Occupation",font=txtfont2,bg='#ffffff')
            field16_window = my_canvas.create_window(966, 660, anchor="nw", window=field16)
            entry16=Entry(window3,width=20,font=txtfont3)
            entry16_window = my_canvas.create_window(970, 700, anchor="nw", window=entry16)

            field17=Label(text="Class",font=txtfont2,bg='#ffffff')
            field17_window = my_canvas.create_window(1266, 160, anchor="nw", window=field17)
            entry17=Entry(window3,width=20,font=txtfont3)
            entry17_window = my_canvas.create_window(1270, 200, anchor="nw", window=entry17)

            field19=Label(text="Previous School",font=txtfont2,bg='#ffffff')
            field19_window = my_canvas.create_window(1266, 260, anchor="nw", window=field19)
            entry19=Entry(window3,width=20,font=txtfont3)
            entry19_window = my_canvas.create_window(1270, 300, anchor="nw", window=entry19)

            field21=Label(text="Result",font=txtfont2,bg='#ffffff')
            field21_window = my_canvas.create_window(1266, 360, anchor="nw", window=field21)
            menu1=OptionMenu(window3,variable1,*OptionList1)
            menu1.config(font=txtfont,bg='#ffffff',width=17,anchor='w')
            menu_window1 = my_canvas.create_window(1270, 400, anchor="nw", window=menu1)

            field20=Label(text="Admission Fee",font=txtfont2,bg='#ffffff')
            field20_window = my_canvas.create_window(1266, 460, anchor="nw", window=field20)
            entry20=Lotfi(window3,width=20,font=txtfont3)
            entry20_window = my_canvas.create_window(1270, 500, anchor="nw", window=entry20)
            

            #field22=Label(text="Uniform",font=txtfont2,bg='#ffffff')
            #field22_window = my_canvas.create_window(1266, 460, anchor="nw", window=field22)
            #entry22=Lotfi(window3,width=20,font=txtfont3)
            #entry22_window = my_canvas.create_window(1270, 500, anchor="nw", window=entry22)

            #field23=Label(text="Other",font=txtfont2,bg='#ffffff')
            #field23_window = my_canvas.create_window(1266, 560, anchor="nw", window=field23)
            #entry23=Lotfi(window3,width=20,font=txtfont3)
            #entry23_window = my_canvas.create_window(1270, 600, anchor="nw", window=entry23)
            
            #field24=Label(text="Description*",font=txtfont2,bg='#ffffff')
            #field24_window = my_canvas.create_window(1266, 660, anchor="nw", window=field24)
            #entry24=Entry(window3,width=20,font=txtfont3)
            #entry24_window = my_canvas.create_window(1270, 700, anchor="nw", window=entry24)

            button4 = Button(window3, text="SAVE", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=to_excel)
            button4.config(height = 2, width=20)
            button4_window = my_canvas.create_window(1270, 600, anchor="nw", window=button4)

        elif state == 3:
            OptionList={'January':'E','February':'F','March':'G','April':'H','May':'I','June':'J','July':'K','August':'L','September':'M','October':'N','November':'O','December':'P'}
            variable = StringVar(window3)
            variable.set("Month")
            button5 = Button(window3, text="Student List", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state1)
            button5.config(height = 2, width=30)
            button5_window = my_canvas.create_window(2, 298, anchor="nw", window=button5)
            button1 = Button(window3, text="New Admission", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state2)
            button1.config(height = 2, width=30)
            button1_window = my_canvas.create_window(2, 338, anchor="nw", window=button1)
            button6 = Button(window3, text="Renew Admission", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=state5)
            button6.config(height = 2, width=30)
            button6_window = my_canvas.create_window(2, 378, anchor="nw", window=button6)
            button2 = Button(window3, text="Fee Entry", bg='#ffffff', highlightthickness=0, bd=0, fg='#3f51cd',font=txtfont)
            button2.config(height = 2, width=30)
            button2_window = my_canvas.create_window(2, 418, anchor="nw", window=button2)
            button3 = Button(window3, text="Entry Log", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state4)
            button3.config(height = 2, width=30)
            button3_window = my_canvas.create_window(2, 458, anchor="nw", window=button3)

            section1=Label(text="Fee Entry",font=txtfont1,bg='#ffffff')
            section1_window = my_canvas.create_window(350, 100, anchor="nw", window=section1)

            field1=Label(text="Name",font=txtfont2,bg='#ffffff')
            field1_window = my_canvas.create_window(366, 180, anchor="nw", window=field1)
            entry1=Entry(width=20,font=txtfont3)
            entry1_window = my_canvas.create_window(370, 220, anchor="nw", window=entry1)

            field2=Label(text="Class",font=txtfont2,bg='#ffffff')
            field2_window = my_canvas.create_window(366, 280, anchor="nw", window=field2)
            entry2=Entry(width=20,font=txtfont3)
            entry2_window = my_canvas.create_window(370, 320, anchor="nw", window=entry2)

            field3=Label(text="Roll No.",font=txtfont2,bg='#ffffff')
            field3_window = my_canvas.create_window(366, 380, anchor="nw", window=field3)
            entry3=Lotfi(width=20,font=txtfont3)
            entry3_window = my_canvas.create_window(370, 420, anchor="nw", window=entry3)


            field9=Label(text="Paying For",font=txtfont2,bg='#ffffff')
            field9_window = my_canvas.create_window(666, 180, anchor="nw", window=field9)

            menu=OptionMenu(window3,variable,*OptionList)
            menu.config(font=txtfont,bg='#ffffff',width=17,anchor='w')
            menu_window = my_canvas.create_window(670, 220, anchor="nw", window=menu)
            
            button3 = Button(window3, text="Add", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=addmonth)
            button3.config(height = 2, width=8)
            button3_window = my_canvas.create_window(840, 216, anchor="nw", window=button3)
            
            field4=Label(text="Fee Amount",font=txtfont2,bg='#ffffff')
            field4_window = my_canvas.create_window(666, 280, anchor="nw", window=field4)
            entry4=Lotfi(width=20,font=txtfont3)
            entry4_window = my_canvas.create_window(670, 320, anchor="nw", window=entry4)
            

            field8=Label(text="Uniform",font=txtfont2,bg='#ffffff')
            field8_window = my_canvas.create_window(666, 380, anchor="nw", window=field8)
            entry8=Lotfi(width=20,font=txtfont3)
            entry8_window = my_canvas.create_window(670, 420, anchor="nw", window=entry8)
            
            field6=Label(text="Other",font=txtfont2,bg='#ffffff')
            field6_window = my_canvas.create_window(966, 180, anchor="nw", window=field6)
            entry6=Lotfi(window3,width=20,font=txtfont3)
            entry6_window = my_canvas.create_window(970, 220, anchor="nw", window=entry6)

            field5=Label(text="Books And Stationaries",font=txtfont2,bg='#ffffff')
            field5_window = my_canvas.create_window(966, 280, anchor="nw", window=field5)
            entry5=Lotfi(width=20,font=txtfont3)
            entry5_window = my_canvas.create_window(970, 320, anchor="nw", window=entry5)


            
            field7=Label(text="Description*",font=txtfont2,bg='#ffffff')
            field7_window = my_canvas.create_window(966, 380, anchor="nw", window=field7)
            entry7=Entry(window3,width=20,font=txtfont3)
            entry7_window = my_canvas.create_window(970, 420, anchor="nw", window=entry7)

            section2=Label(text="For the month of",textvariable=u,font=txtfont,bg='#ffffff',fg="grey")
            section2_window = my_canvas.create_window(368, 450, anchor="nw", window=section2)
            
            button4 = Button(window3, text="SAVE", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=to_excel1)
            button4.config(height = 2, width=20)
            button4_window = my_canvas.create_window(370, 500, anchor="nw", window=button4)
            
        elif state == 1:
            global count
            global tags
            tags=[]
            count=0
            OptionList={"Fee Status":23,"AdmissionID":24,"Class":18,"Roll No.":19,"Batch":0,"First Name":0,"Last name":1,"Gender":2,"Age":3,"Community":14,"Tribe/Caste":15,"Religion":8,"Father's Name":4,"Mother's Name":5,"Gaurdian's Name":6,"DOB":7,"Address":9,"Father's Contact No.":10,"Mother's Contact No.":11,"Identity Mark":12,"Blood Group":13,"Prev. School":20,"Result":21}
            variable = StringVar(window3)
            variable.set("Select Filter")
            button5 = Button(window3, text="Student List", bg='#ffffff', highlightthickness=0, bd=0, fg='#3f51cd',font=txtfont)
            button5.config(height = 2, width=30)
            button5_window = my_canvas.create_window(2, 298, anchor="nw", window=button5)
            button1 = Button(window3, text="New Admission", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=state2)
            button1.config(height = 2, width=30)
            button1_window = my_canvas.create_window(2, 338, anchor="nw", window=button1)
            button6 = Button(window3, text="Renew Admission", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=state5)
            button6.config(height = 2, width=30)
            button6_window = my_canvas.create_window(2, 378, anchor="nw", window=button6)
            button2 = Button(window3, text="Fee Entry", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state3)
            button2.config(height = 2, width=30)
            button2_window = my_canvas.create_window(2, 418, anchor="nw", window=button2)
            button3 = Button(window3, text="Entry Log", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state4)
            button3.config(height = 2, width=30)
            button3_window = my_canvas.create_window(2, 458, anchor="nw", window=button3)
            
            section1=Label(text="Student List",font=txtfont1,bg='#ffffff')
            section1_window = my_canvas.create_window(350, 100, anchor="nw", window=section1)

            section2=Label(text="Tags: Students",textvariable=v,font=txtfont,bg='#ffffff',fg="grey")
            section2_window = my_canvas.create_window(245, 275, anchor="nw", window=section2)

            section3=Label(text="Count:",textvariable=v1,font=txtfont,bg='#ffffff',fg="grey")
            section3_window = my_canvas.create_window(1400, 275, anchor="nw", window=section3)
            
            menu=OptionMenu(window3,variable,*OptionList)
            menu.config(font=txtfont,bg='#ffffff',width=17,anchor='w')
            menu_window = my_canvas.create_window(355, 180, anchor="nw", window=menu)

            c1 = tk.Checkbutton(window3, text='Reverse',variable=var1, onvalue=1, offvalue=0,bg='#ffffff')
            check_window = my_canvas.create_window(355, 220, anchor="nw", window=c1)

            entry1=Entry(window3,width=11,font=txtfont4)
            entry1_window = my_canvas.create_window(540, 182, anchor="nw", window=entry1)        

            button4 = Button(window3, text="Add Filter", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=rd_excel)
            button4.config(height = 2, width=10)
            button4_window = my_canvas.create_window(680, 177, anchor="nw", window=button4)

            button5 = Button(window3, text="Remove Filter", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=remove_filter)
            button5.config(height = 2, width=12)
            button5_window = my_canvas.create_window(770, 177, anchor="nw", window=button5)

            
            tree_frame = Frame(window3,width=1289, height=500)
            tree_frame_window=my_canvas.create_window(245, 298, anchor="nw", window=tree_frame)
            tree_scrolly = Scrollbar(tree_frame)
            tree_scrolly.pack(side=RIGHT, fill=Y)
            tree_frame.pack_propagate(0)
            treeview=ttk.Treeview(tree_frame, yscrollcommand=tree_scrolly.set, selectmode="extended")
            treeview.pack(side="top", fill="both", expand=True)
            scroll = Scrollbar(tree_frame, orient=HORIZONTAL, command=treeview.xview)
            treeview.configure(xscrollcommand=scroll.set)
            tree_scrolly.config(command=treeview.yview)
            scroll.pack(side="bottom", fill="x")
            treeview["column"] = list(df.columns)
            treeview["show"] = "headings"
            treeview.tag_configure('oddrow', background="white")
            treeview.tag_configure('evenrow', background="#F7F7F7")
            for column in treeview["column"]:
                treeview.heading(column, text=column,anchor="w",command=lambda _column=column: \
                                                 treeview_sort_column(treeview, _column, False))
            for row in df_rows:
                if count % 2 == 0:
                    treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('evenrow',))
                else:
                    treeview.insert(parent='', index='end', iid=count, text="", values=row, tags=('oddrow',))
                count=count+1
            v1.set("Count: "+str(count))
            treeview.bind("<Double-1>", OnDoubleClick)
        elif state == 4:
            OptionList=["Admissions","Fee Entries","Fee Status","Total"]
            variable = StringVar(window3)
            variable.set("Total")
            OptionList1={"Batch":0,"Class":5,"AdmissionID":2}
            variable1 = StringVar(window3)
            variable1.set("Select Filter")
            button5 = Button(window3, text="Student List", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state1)
            button5.config(height = 2, width=30)
            button5_window = my_canvas.create_window(2, 298, anchor="nw", window=button5)
            button1 = Button(window3, text="New Admission", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=state2)
            button1.config(height = 2, width=30)
            button1_window = my_canvas.create_window(2, 338, anchor="nw", window=button1)
            button6 = Button(window3, text="Renew Admission", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=state5)
            button6.config(height = 2, width=30)
            button6_window = my_canvas.create_window(2, 378, anchor="nw", window=button6)
            button2 = Button(window3, text="Fee Entry", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state3)
            button2.config(height = 2, width=30)
            button2_window = my_canvas.create_window(2, 418, anchor="nw", window=button2)
            button3 = Button(window3, text="Entry Log", bg='#ffffff', highlightthickness=0, bd=0, fg='#3f51cd',font=txtfont)
            button3.config(height = 2, width=30)
            button3_window = my_canvas.create_window(2, 458, anchor="nw", window=button3)

            section1=Label(text="Entry Log",font=txtfont1,bg='#ffffff')
            section1_window = my_canvas.create_window(350, 100, anchor="nw", window=section1)

            section2=Label(text="Tags: Entries",textvariable=v2,font=txtfont,bg='#ffffff',fg="grey")
            section2_window = my_canvas.create_window(245, 275, anchor="nw", window=section2)

            section3=Label(text="Count :",textvariable=v1,font=txtfont,bg='#ffffff',fg="grey")
            section3_window = my_canvas.create_window(1400, 275, anchor="nw", window=section3)

            field1=Label(text="Date (dd/mm/yy)",font=txtfont,bg='#ffffff',fg="grey")
            field1_window = my_canvas.create_window(355, 156, anchor="nw", window=field1)
            entry1=Entry(window3,width=11,font=txtfont4)
            entry1_window = my_canvas.create_window(355, 180, anchor="nw", window=entry1)
            field2=Label(text="To",font=txtfont2,bg='#ffffff')
            field2_window = my_canvas.create_window(497, 180, anchor="nw", window=field2)
            entry2=Entry(window3,width=11,font=txtfont4)
            entry2_window = my_canvas.create_window(540, 180, anchor="nw", window=entry2)

            button4 = Button(window3, text="Search", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=log_call)
            button4.config(height = 2, width=10)
            button4_window = my_canvas.create_window(680, 177, anchor="nw", window=button4)

            menu=OptionMenu(window3,variable,*OptionList)
            menu.config(font=txtfont,bg='#ffffff',width=12,anchor='w')
            menu_window = my_canvas.create_window(354, 220, anchor="nw", window=menu)

            menu1=OptionMenu(window3,variable1,*OptionList1)
            menu1.config(font=txtfont,bg='#ffffff',width=17,anchor='w')
            menu1_window = my_canvas.create_window(905, 180, anchor="nw", window=menu1)

            c1 = tk.Checkbutton(window3, text='Reverse',variable=var1, onvalue=1, offvalue=0, bg='#ffffff')
            check_window = my_canvas.create_window(905, 220, anchor="nw", window=c1)

            entry3=Entry(window3,width=11,font=txtfont4)
            entry3_window = my_canvas.create_window(1090, 182, anchor="nw", window=entry3)        

            button6 = Button(window3, text="Add Filter", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=excel_read)
            button6.config(height = 2, width=10)
            button6_window = my_canvas.create_window(1230, 177, anchor="nw", window=button6)

            button5 = Button(window3, text="Remove Filter", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=remove_filter1)
            button5.config(height = 2, width=12)
            button5_window = my_canvas.create_window(1320, 177, anchor="nw", window=button5)
            
            tree_frame = Frame(window3,width=1289, height=500)
            tree_frame_window=my_canvas.create_window(245, 298, anchor="nw", window=tree_frame)
            tree_scrolly = Scrollbar(tree_frame)
            tree_scrolly.pack(side=RIGHT, fill=Y)
            tree_frame.pack_propagate(0)
            treeview=ttk.Treeview(tree_frame, yscrollcommand=tree_scrolly.set, selectmode="extended")
            treeview.pack(side="top", fill="both", expand=True)
            scroll = Scrollbar(tree_frame, orient=HORIZONTAL, command=treeview.xview)
            treeview.configure(xscrollcommand=scroll.set)
            tree_scrolly.config(command=treeview.yview)
            scroll.pack(side="bottom", fill="x")
            log(1)
            
        elif state == 5:
            OptionList=['Pass','Fail']
            variable = StringVar(window3)
            variable.set("Pass")
            button5 = Button(window3, text="Student List", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state1)
            button5.config(height = 2, width=30)
            button5_window = my_canvas.create_window(2, 298, anchor="nw", window=button5)
            button1 = Button(window3, text="New Admission", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=state2)
            button1.config(height = 2, width=30)
            button1_window = my_canvas.create_window(2, 338, anchor="nw", window=button1)
            button6 = Button(window3, text="Renew Admission", bg='#ffffff', highlightthickness=0, bd=0, fg='#3f51cd',font=txtfont)
            button6.config(height = 2, width=30)
            button6_window = my_canvas.create_window(2, 378, anchor="nw", window=button6)
            button2 = Button(window3, text="Fee Entry", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont, command=state3)
            button2.config(height = 2, width=30)
            button2_window = my_canvas.create_window(2, 418, anchor="nw", window=button2)
            button3 = Button(window3, text="Entry Log", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=state4)
            button3.config(height = 2, width=30)
            button3_window = my_canvas.create_window(2, 458, anchor="nw", window=button3)


            section1=Label(text="Renew Admission",font=txtfont1,bg='#ffffff')
            section1_window = my_canvas.create_window(350, 100, anchor="nw", window=section1)

            field1=Label(text="First Name",font=txtfont2,bg='#ffffff')
            field1_window = my_canvas.create_window(366, 180, anchor="nw", window=field1)
            entry1=Entry(width=20,font=txtfont3)
            entry1_window = my_canvas.create_window(370, 220, anchor="nw", window=entry1)

            field2=Label(text="Last Name",font=txtfont2,bg='#ffffff')
            field2_window = my_canvas.create_window(366, 280, anchor="nw", window=field2)
            entry2=Entry(width=20,font=txtfont3)
            entry2_window = my_canvas.create_window(370, 320, anchor="nw", window=entry2)

            field3=Label(text="Class",font=txtfont2,bg='#ffffff')
            field3_window = my_canvas.create_window(666, 180, anchor="nw", window=field3)
            entry3=Entry(width=20,font=txtfont3)
            entry3_window = my_canvas.create_window(670, 220, anchor="nw", window=entry3)
            
            field6=Label(text="Result",font=txtfont2,bg='#ffffff')
            field6_window = my_canvas.create_window(666, 280, anchor="nw", window=field6)
            menu=OptionMenu(window3,variable,*OptionList)
            menu.config(font=txtfont,bg='#ffffff',width=17,anchor='w')
            menu_window = my_canvas.create_window(670, 320, anchor="nw", window=menu)        

            field4=Label(text="Admission ID",font=txtfont2,bg='#ffffff')
            field4_window = my_canvas.create_window(966, 180, anchor="nw", window=field4)
            entry4=Entry(width=20,font=txtfont3)
            entry4_window = my_canvas.create_window(970, 220, anchor="nw", window=entry4)

            field5=Label(text="Admission Fee",font=txtfont2,bg='#ffffff')
            field5_window = my_canvas.create_window(966, 280, anchor="nw", window=field5)
            entry5=Lotfi(width=20,font=txtfont3)
            entry5_window = my_canvas.create_window(970, 320, anchor="nw", window=entry5)

            button4 = Button(window3, text="SAVE", bg='#3f51cd', highlightthickness=0, bd=0, fg='#ffffff',font=txtfont,command=to_excel2)
            button4.config(height = 2, width=20)
            button4_window = my_canvas.create_window(370, 380, anchor="nw", window=button4)
    form()
    




def check_cred():
    if entry1.get()!="admin" and entry2.get()!="1234":
        messagebox.showwarning("Oops!","Invalid Credentials!")
        entry1.delete(0,'end')
        entry2.delete(0,'end')
        return
    else:
        window.destroy()
        start()
    return
window = Tk()
window.title("Admin Login")
window.iconbitmap('data/icon.ico')
w = 250 # width for the Tk root
h = 140 # height for the Tk root
ws = window.winfo_screenwidth() # width of the screen
hs = window.winfo_screenheight() # height of the screen
x = (ws/2) - (w/2)
y = (hs/2) - (h/2)
window.geometry('%dx%d+%d+%d' % (w, h, x, y))
window.resizable(width=False, height=False)
my_canvas = Canvas(window, width=285, height=100)
my_canvas.pack(fill="both", expand=True)

field1=Label(text="User Name*")
field1_window = my_canvas.create_window(20, 25, anchor="nw", window=field1)
entry1=Entry(width=20)
entry1_window = my_canvas.create_window(100, 25, anchor="nw", window=entry1)
field2=Label(text="Password*")
field2_window = my_canvas.create_window(20, 60, anchor="nw", window=field2)
entry2=Entry(width=20,show='*')
entry2_window = my_canvas.create_window(100, 60, anchor="nw", window=entry2)
button4 = Button(window, text="Login", command=check_cred)
button4.config(height = 1, width=10)
button4_window = my_canvas.create_window(102, 100, anchor="nw", window=button4)

tags=[]
months=[]
today = date.today()
state=1
substate=0
count=0
index=0
cls=[]
batch=today.strftime("%Y")
df = pd.read_excel('./data/studentlist.xlsx')
df_rows = df.to_numpy().tolist()
for value in df_rows:
    subdate1=value[22].split("/")
    if subdate1[2]!=batch:
        df_rows =list(filter(partial(ne, value), df_rows))
    else:
        continue
 
class Lotfi(tk.Entry):
    def __init__(self, master=None, **kwargs):
        self.var = tk.StringVar()
        tk.Entry.__init__(self, master, textvariable=self.var, **kwargs)
        self.old_value = ''
        self.var.trace('w', self.check)
        self.get, self.set = self.var.get, self.var.set

    def check(self, *args):
        if self.get().isdigit(): 
            # the current value is only digits; allow this
            self.old_value = self.get()
        else:
            # there's non-digit characters in the input; reject this 
            self.set(self.old_value)
